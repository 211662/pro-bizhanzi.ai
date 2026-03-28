import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────────
header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2C416D", end_color="2C416D", fill_type="solid")
done_fill = PatternFill(start_color="D4F7E5", end_color="D4F7E5", fill_type="solid")
done_font = Font(name="Arial", color="0C7054")
notyet_fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
notyet_font = Font(name="Arial", color="AC2929")
partial_fill = PatternFill(start_color="FFF1B8", end_color="FFF1B8", fill_type="solid")
partial_font = Font(name="Arial", color="874D00")
verify_fill = PatternFill(start_color="E3ECF6", end_color="E3ECF6", fill_type="solid")
verify_font = Font(name="Arial", color="42578B")
normal_font = Font(name="Arial", size=10)
thin_border = Border(
    left=Side(style="thin", color="CFCFCF"),
    right=Side(style="thin", color="CFCFCF"),
    top=Side(style="thin", color="CFCFCF"),
    bottom=Side(style="thin", color="CFCFCF"),
)
wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="center")


def style_status(cell, status):
    s = status.lower()
    if s == "done":
        cell.fill, cell.font = done_fill, done_font
    elif s == "not yet":
        cell.fill, cell.font = notyet_fill, notyet_font
    elif s in ("partial", "needs improvement"):
        cell.fill, cell.font = partial_fill, partial_font
    elif s == "need to verify":
        cell.fill, cell.font = verify_fill, verify_font


def setup_sheet(ws):
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 90
    ws.column_dimensions["D"].width = 18


def write_header(ws, row):
    for col, h in enumerate(["STT", "Checklist", "Content", "Status"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        c.border = thin_border
    ws.row_dimensions[row].height = 22
    return row + 1


def write_row(ws, row, stt, checklist, content, status):
    vals = [stt, checklist, content, status]
    for col, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = normal_font
        c.border = thin_border
        c.alignment = wrap_align
    if status:
        style_status(ws.cell(row=row, column=4), status)
    return row + 1


def build_sheet(ws, rows):
    setup_sheet(ws)
    row = write_header(ws, 1)
    for r in rows:
        row = write_row(ws, row, *r)


# ════════════════════════════════════════════════════════════
# SHEET 1: Base
# ════════════════════════════════════════════════════════════
ws_base = wb.active
ws_base.title = "Base"
build_sheet(ws_base, [
    ("1", "Install Google Tag Manager", "GTM ID: GTM-T7DLMWMZ (installed in <head> and <body> noscript)", "Done"),
    ("", "Install Google Analytics", "No standalone GA code found. Likely configured via GTM", "Need to verify"),
    ("", "Install Google Search Console", "No google-site-verification meta tag found. May be verified via DNS or GTM", "Need to verify"),
    ("", "Install Bing Webmaster", "No msvalidate.01 meta tag found", "Not yet"),
    ("2", "Add proper title tags to all pages (unique)", "All pages share the same title — NOT unique per page", "Not yet"),
    ("3", "Add meta descriptions to all pages (unique)", "All pages share the same meta description — NOT unique per page", "Not yet"),
    ("4", "Add canonical URLs (self-referencing)", "No <link rel=canonical> tag found on any page", "Not yet"),
    ("5", "Add Open Graph tags", "og:title, og:url, og:site_name, og:description, og:image, og:type present. fb:app_id empty. og:locale missing", "Partial"),
    ("6", "Add Twitter Card tags", "No twitter:card, twitter:title, twitter:description, twitter:image found", "Not yet"),
    ("7", "Add basic structured data (WebSite, Organization)", "WebSite schema with SearchAction present. Organization schema missing", "Partial"),
    ("8", "Heading", "SPA (Angular) — headings rendered client-side, not in initial HTML source", "Need to verify"),
    ("9", "Schema", "WebSite + SearchAction only. No Organization, SoftwareApplication, FAQ schema", "Partial"),
    ("10", "Optimize all images (alt text, compression)", "SPA — images loaded dynamically. Alt text audit needed on rendered pages", "Not yet"),
    ("11", "Add FAQ schema", "No FAQ structured data found", "Not yet"),
    ("12", "Add hreflang tags for multilingual", "Site supports vi, en, ja, ko, ru via ?hl= param. No <link rel=alternate hreflang> tags found", "Not yet"),
    ("13", "Review and improve internal linking", "Footer: Introduction, Privacy, Terms, Help, Guides. Nav: Home, Translate, Test, Notebook, Community, Upgrade", "Need to verify"),
    ("14", "Add related content sections", "Topics section exists (HSK vocab, themed topics). Community with user contributions", "Partial"),
    ("15", "Create privacy policy page", "Available at /other/privacy-policy", "Done"),
    ("16", "Create terms of service page", "Available at /other/term", "Done"),
    ("17", "Social media channels", "facebook.com/tudientrungviet.bizhan\ninstagram.com/bizhan.chinesedict_\ntiktok.com/@bizhan.official\nzalo.me/0976696764", "Done"),
])

# ════════════════════════════════════════════════════════════
# SHEET 2: Homepage
# ════════════════════════════════════════════════════════════
ws = wb.create_sheet("Homepage")
build_sheet(ws, [
    ("1", "URL", "https://bizhan.net/", ""),
    ("2", "Title tag", "Chinese-English, Chinese-Vietnamese Online Dictionary (39 chars) — should include brand BizHan", "Needs improvement"),
    ("3", "Suggested title", "BizHan — Free Chinese-Vietnamese & Chinese-English Dictionary Online", ""),
    ("4", "Meta description", "BizHan free online Chinese-Vietnamese, Chinese-English dictionary. Look up characters by radical, stroke, image. Complete grammar, examples and illustrations (155 chars)", "Done"),
    ("5", "Canonical URL", "Missing — should be https://bizhan.net/", "Not yet"),
    ("6", "H1 tag", "SPA rendered — not in source HTML", "Need to verify"),
    ("7", "OG title", "Chinese-English, Chinese-Vietnamese Online Dictionary", "Needs improvement"),
    ("8", "OG url", "https://bizhan.net", "Done"),
    ("9", "OG image", "ic_bizhan1024x1024.png", "Done"),
    ("10", "OG type", "website", "Done"),
    ("11", "OG description", "Same as meta description", "Done"),
    ("12", "OG site_name", "BizHan — Chinese-English, Chinese-Vietnamese Online Dictionary", "Done"),
    ("13", "OG locale", "Missing", "Not yet"),
    ("14", "fb:app_id", "Empty value", "Not yet"),
    ("15", "Twitter Card", "Missing — no twitter:card, twitter:title, twitter:description, twitter:image", "Not yet"),
    ("16", "Schema markup", "WebSite + SearchAction ✓ | Organization: missing | SoftwareApplication: missing", "Partial"),
    ("17", "Keywords meta", "Long keyword list present (Vietnamese + Chinese learning related)", "Done"),
    ("18", "Robots meta", "index,follow", "Done"),
    ("19", "Language meta", '<html lang="vi"> + <meta name="language" content="vi">', "Done"),
    ("20", "Hreflang tags", "Supports vi, en, ja, ko, ru via ?hl= param. No hreflang tags", "Not yet"),
    ("21", "Favicon", "assets/images/ic_logo.ico", "Done"),
    ("22", "Manifest (PWA)", "manifest.json linked", "Done"),
    ("23", "DMCA protection", "DMCA badge script + verification meta tag present", "Done"),
])

# ════════════════════════════════════════════════════════════
# SUB-PAGE SHEETS (1 sheet per page)
# ════════════════════════════════════════════════════════════
sub_pages = [
    ("Translate",       "/translate",                           "BizHan Translate — Chinese Translation Tool | BizHan",             "Chinese text and document translation tool. Translate between Chinese, Vietnamese, English, Japanese, Korean and more."),
    ("Test",            "/test",                                "HSK Mock Exam — Practice Chinese Tests Online | BizHan",           "Practice HSK mock exams online for free. Test your Chinese proficiency from HSK 1 to HSK 6."),
    ("Notebook",        "/notebook",                            "Vocabulary Notebook — Save & Review Chinese Words | BizHan",       "Save and organize Chinese vocabulary. Review HSK and TOCFL word lists with smart flashcard system."),
    ("Community",       "/community",                           "Community — Learn Chinese Together | BizHan",                      "Join the BizHan community. Share Chinese learning tips, Hanzi mnemonics, and connect with learners."),
    ("Upgrade",         "/upgrade",                             "Upgrade to Premium — BizHan Pro Plans",                            "Unlock all BizHan premium features. Ad-free, unlimited translations, advanced AI tools and more."),
    ("Introduction",    "/other/introduction",                  "About BizHan — Chinese-Vietnamese Dictionary",                     "Learn about BizHan, the leading Chinese-Vietnamese and Chinese-English online dictionary with AI features."),
    ("Privacy Policy",  "/other/privacy-policy",                "Privacy Policy | BizHan",                                          "BizHan's privacy policy. How we collect, use and protect your personal data."),
    ("Terms",           "/other/term",                          "Terms of Service | BizHan",                                        "BizHan's terms of service. Rules and conditions for using the dictionary and learning platform."),
    ("Help",            "/other/help",                          "Help Center | BizHan",                                             "Get help with BizHan features. FAQs, guides and support for the Chinese dictionary and tools."),
    ("HSK 1",           "/notebook/detail/HSK-1",               "HSK 1 Vocabulary List — 150 Basic Chinese Words | BizHan",         "Complete HSK 1 vocabulary list with pinyin, meaning and examples. 150 essential words for beginners."),
    ("HSK 2",           "/notebook/detail/HSK-2",               "HSK 2 Vocabulary List — 300 Chinese Words | BizHan",               "Complete HSK 2 vocabulary list with pinyin, meaning and examples. 300 elementary Chinese words."),
    ("HSK 3",           "/notebook/detail/HSK-3",               "HSK 3 Vocabulary List — 600 Chinese Words | BizHan",               "Complete HSK 3 vocabulary list with pinyin, meaning and examples. 600 intermediate Chinese words."),
    ("HSK 4",           "/notebook/detail/HSK-4",               "HSK 4 Vocabulary List — 1200 Chinese Words | BizHan",              "Complete HSK 4 vocabulary list with pinyin, meaning and examples. 1200 upper-intermediate words."),
    ("HSK 5",           "/notebook/detail/HSK-5",               "HSK 5 Vocabulary List — 2500 Chinese Words | BizHan",              "Complete HSK 5 vocabulary list with pinyin, meaning and examples. 2500 advanced Chinese words."),
    ("HSK 6",           "/notebook/detail/HSK-6",               "HSK 6 Vocabulary List — 5000 Chinese Words | BizHan",              "Complete HSK 6 vocabulary list with pinyin, meaning and examples. 5000 full proficiency words."),
    ("TOCFL 1",         "/notebook/detail/TOCFL-1",             "TOCFL Band A Level 1 Vocabulary | BizHan",                         "TOCFL Band A Level 1 vocabulary. Essential Traditional Chinese words for Taiwan proficiency test."),
    ("TOCFL 2",         "/notebook/detail/TOCFL-2",             "TOCFL Band A Level 2 Vocabulary | BizHan",                         "TOCFL Band A Level 2 vocabulary for Taiwan Chinese proficiency test preparation."),
    ("TOCFL 3",         "/notebook/detail/TOCFL-3",             "TOCFL Band B Level 3 Vocabulary | BizHan",                         "TOCFL Band B Level 3 vocabulary for intermediate Taiwan Chinese proficiency."),
    ("TOCFL 4",         "/notebook/detail/TOCFL-4",             "TOCFL Band B Level 4 Vocabulary | BizHan",                         "TOCFL Band B Level 4 vocabulary for upper-intermediate Taiwan Chinese proficiency."),
    ("TOCFL 5",         "/notebook/detail/TOCFL-5",             "TOCFL Band C Level 5 Vocabulary | BizHan",                         "TOCFL Band C Level 5 vocabulary for advanced Taiwan Chinese proficiency test."),
    ("Grammar A1",      "/search/grammar/A1",                   "Chinese Grammar A1 — Beginner Structures | BizHan",                "Beginner Chinese grammar A1 structures with clear explanations and example sentences."),
    ("Grammar A2",      "/search/grammar/A2",                   "Chinese Grammar A2 — Elementary Structures | BizHan",              "Elementary Chinese grammar A2 structures with explanations and examples."),
    ("Grammar B1",      "/search/grammar/B1",                   "Chinese Grammar B1 — Intermediate Structures | BizHan",            "Intermediate Chinese grammar B1 structures with explanations and examples."),
    ("Grammar B2",      "/search/grammar/B2",                   "Chinese Grammar B2 — Upper-Intermediate Structures | BizHan",      "Upper-intermediate Chinese grammar B2 structures with explanations and examples."),
    ("Grammar C1",      "/search/grammar/C1",                   "Chinese Grammar C1 — Advanced Structures | BizHan",                "Advanced Chinese grammar C1 structures with detailed explanations and examples."),
    ("Grammar C2",      "/search/grammar/C2",                   "Chinese Grammar C2 — Proficiency Structures | BizHan",             "Proficiency-level Chinese grammar C2 structures with comprehensive explanations."),
    ("Guide Payment",   "/guide/online-payment",                "Online Payment Guide | BizHan",                                    "Step-by-step guide for online payment on BizHan. Supported methods and instructions."),
    ("Payment Policy",  "/guide/payment-policy",                "Payment Policy | BizHan",                                          "BizHan payment policy. Accepted methods, billing and transaction terms."),
    ("Check Goods",     "/guide/check-goods",                   "Product Inspection Policy | BizHan",                               "BizHan's product inspection and verification policy."),
    ("Info Security",   "/guide/personal-information-security",  "Personal Information Security Policy | BizHan",                   "How BizHan protects your personal information and data security measures."),
    ("Shipping",        "/guide/shipping-receiving-policy",      "Shipping & Receiving Policy | BizHan",                            "BizHan shipping and delivery policy. Processing times and delivery methods."),
    ("Complaints",      "/guide/resolve-complaints",             "Complaints Resolution Process | BizHan",                          "How to file and resolve complaints with BizHan. Process and response times."),
    ("Refund",          "/guide/exchange-refund",                "Exchange & Refund Policy | BizHan",                                "BizHan exchange and refund policy. Conditions, timeframes and refund requests."),
]

for sheet_name, path, suggested_title, suggested_desc in sub_pages:
    url = f"https://bizhan.net{path}"
    ws = wb.create_sheet(sheet_name)
    build_sheet(ws, [
        ("1", "URL", url, ""),
        ("2", "Current title tag", "Chinese-English, Chinese-Vietnamese Online Dictionary — same as all pages, NOT unique", "Not yet"),
        ("3", "Suggested title tag", suggested_title, ""),
        ("4", "Current meta description", "Same as homepage — NOT unique", "Not yet"),
        ("5", "Suggested meta description", suggested_desc, ""),
        ("6", "Canonical URL", f"Missing — should be {url}", "Not yet"),
        ("7", "OG title", "Same as homepage — NOT page-specific", "Not yet"),
        ("8", "OG description", "Same as homepage — NOT page-specific", "Not yet"),
        ("9", "OG url", f"Should be {url}", "Not yet"),
        ("10", "OG image", "Should have page-specific image", "Not yet"),
        ("11", "Twitter Card", "Missing", "Not yet"),
        ("12", "H1 tag", "SPA rendered — need to verify in browser", "Need to verify"),
        ("13", "Schema markup", "No page-specific schema", "Not yet"),
        ("14", "Hreflang tags", f"Missing — should link vi/en/ja/ko/ru versions of {path}", "Not yet"),
        ("15", "Internal links", "Need to audit inbound/outbound internal links", "Need to verify"),
    ])

# ════════════════════════════════════════════════════════════
output = "/Users/linh/Desktop/github/pro-bizhanzi.ai/Onpage_hanzii.xlsx"
wb.save(output)
print(f"✅ Created: {output}")
print(f"   Total sheets: {len(wb.sheetnames)}")
for i, name in enumerate(wb.sheetnames):
    print(f"   {i+1:2d}. {name}")
